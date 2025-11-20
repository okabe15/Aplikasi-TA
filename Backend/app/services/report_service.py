"""
Complete Report Service - PYTHON 3.13 COMPATIBLE VERSION
Fixed CACHE + JUMP_BACKWARD error
"""

import httpx
from app.config import settings
from typing import List, Dict, Optional, Any
import openpyxl
from datetime import datetime, timedelta
from pony.orm import db_session, select, desc, count
from app.database.models import User, UserProgress, LearningModule, Exercise, UserAnswer
import pandas as pd
from io import BytesIO
import json

# ReportLab imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT  # ✅ ADD TA_LEFT

# OpenPyXL imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # ✅ ADD Border, Side
from openpyxl.utils import get_column_letter

import logging
logger = logging.getLogger(__name__)

class ReportService:
    """Service for generating various reports"""
    
    REPORT_TYPES = {
        'student_progress': 'Individual Student Progress Report',
        'class_overview': 'Class Overview Report', 
        'module_performance': 'Module Performance Report',
        'exercise_analysis': 'Exercise Analysis Report',
        'comparative_analysis': 'Comparative Analysis Report',
        'engagement_metrics': 'Student Engagement Report',
        'achievement_summary': 'Achievement Summary Report',
        'weekly_summary': 'Weekly Summary Report'
    }
    
    @db_session
    def generate_student_progress_report(self, student_id: int, format: str = 'pdf') -> bytes:
        """Generate detailed progress report for a specific student"""
        
        student = User.get(id=student_id)
        if not student:
            raise ValueError("Student not found")
        
        # ✅ FIX: Convert to list immediately to avoid loop issues
        all_progress = list(student.progress_records)
        
        total_modules = len(all_progress)
        completed_modules = sum(1 for p in all_progress if p.completed)
        total_score = sum(p.total_score for p in all_progress)
        total_questions = sum(p.total_questions for p in all_progress)
        correct_answers = sum(p.correct_answers for p in all_progress)
        accuracy = (correct_answers / total_questions * 100) if total_questions > 0 else 0
        
        # Get detailed progress by module
        module_details = []
        for prog in all_progress:
            module_details.append({
                'module_id': prog.module.id,
                'module_title': prog.module.classic_text[:50] + "...",
                'started_at': prog.started_at,
                'completed_at': prog.completed_at,
                'score': prog.total_score,
                'questions': prog.total_questions,
                'correct': prog.correct_answers,
                'accuracy': (prog.correct_answers / prog.total_questions * 100) if prog.total_questions > 0 else 0,
                'completed': prog.completed
            })
        
        # Get exercise performance by type
        exercise_performance = self._get_exercise_performance(student_id)
        
        # Get time-based progress
        weekly_progress = self._get_weekly_progress(student_id)
        
        if format == 'pdf':
            return self._generate_pdf_report(
                'student_progress',
                {
                    'student': student,
                    'total_modules': total_modules,
                    'completed_modules': completed_modules,
                    'total_score': total_score,
                    'accuracy': accuracy,
                    'module_details': module_details,
                    'exercise_performance': exercise_performance,
                    'weekly_progress': weekly_progress
                }
            )
        elif format == 'excel':
            return self._generate_excel_report(
                'student_progress',
                {
                    'student': student,
                    'module_details': module_details,
                    'exercise_performance': exercise_performance,
                    'weekly_progress': weekly_progress
                }
            )
        else:
            return json.dumps({
                'student': {
                    'id': student.id,
                    'name': student.full_name,
                    'username': student.username
                },
                'summary': {
                    'total_modules': total_modules,
                    'completed_modules': completed_modules,
                    'total_score': total_score,
                    'accuracy': accuracy
                },
                'module_details': module_details,
                'exercise_performance': exercise_performance,
                'weekly_progress': weekly_progress
            }, default=str).encode()
    
    @db_session
    def generate_class_overview_report(self, format: str = 'pdf') -> bytes:
        """Generate overview report for all students"""
        
        # ✅ FIX: Get all students as list first
        all_students = list(User.select(lambda u: u.role == 'student'))
        
        class_data = []
        for student in all_students:
            # ✅ FIX: Convert to list immediately
            progress_list = list(student.progress_records)
            
            total_score = sum(p.total_score for p in progress_list)
            modules_completed = sum(1 for p in progress_list if p.completed)
            total_questions = sum(p.total_questions for p in progress_list)
            correct_answers = sum(p.correct_answers for p in progress_list)
            accuracy = (correct_answers / total_questions * 100) if total_questions > 0 else 0
            
            # Get latest activity
            latest_activity = None
            if progress_list:
                dates = [p.completed_at if p.completed_at else p.started_at for p in progress_list]
                dates = [d for d in dates if d is not None]
                if dates:
                    latest_activity = max(dates)
            
            class_data.append({
                'student_id': student.id,
                'student_name': student.full_name,
                'username': student.username,
                'total_score': total_score,
                'modules_completed': modules_completed,
                'accuracy': accuracy,
                'latest_activity': latest_activity,
                'total_questions': total_questions,
                'correct_answers': correct_answers
            })
        
        # Sort by total score
        class_data.sort(key=lambda x: x['total_score'], reverse=True)
        
        # Calculate class statistics
        class_stats = {
            'total_students': len(all_students),
            'active_students': sum(1 for s in class_data if s['latest_activity']),
            'avg_score': sum(s['total_score'] for s in class_data) / len(class_data) if class_data else 0,
            'avg_accuracy': sum(s['accuracy'] for s in class_data) / len(class_data) if class_data else 0,
            'top_performers': class_data[:5] if len(class_data) >= 5 else class_data
        }
        
        if format == 'pdf':
            return self._generate_pdf_report('class_overview', {
                'class_data': class_data,
                'class_stats': class_stats
            })
        elif format == 'excel':
            return self._generate_excel_report('class_overview', {
                'class_data': class_data,
                'class_stats': class_stats
            })
        else:
            return json.dumps({
                'class_data': class_data,
                'statistics': class_stats
            }, default=str).encode()
    
    @db_session
    def generate_module_performance_report(self, module_id: Optional[str] = None, format: str = 'pdf') -> bytes:
        """Generate performance report for specific module or all modules"""
        
        if module_id:
            modules_list = [LearningModule.get(id=module_id)]
            if not modules_list[0]:
                raise ValueError("Module not found")
        else:
            # ✅ FIX: Convert to list immediately
            modules_list = list(LearningModule.select())
        
        module_data = []
        for module in modules_list:
            # ✅ FIX: Convert to list
            progress_list = list(module.user_progress)
            exercises_list = list(module.exercises)
            
            # Calculate module statistics
            total_attempts = len(progress_list)
            completed_attempts = sum(1 for p in progress_list if p.completed)
            
            avg_score = sum(p.total_score for p in progress_list) / total_attempts if total_attempts > 0 else 0
            
            accuracy_sum = 0
            for p in progress_list:
                if p.total_questions > 0:
                    accuracy_sum += (p.correct_answers / p.total_questions * 100)
            avg_accuracy = accuracy_sum / total_attempts if total_attempts > 0 else 0
            
            # Get exercise difficulty analysis
            exercise_stats = []
            for exercise in exercises_list:
                answers_list = list(exercise.user_answers)
                
                correct_count = sum(1 for a in answers_list if a.is_correct)
                total_count = len(answers_list)
                success_rate = (correct_count / total_count * 100) if total_count > 0 else 0
                
                exercise_stats.append({
                    'exercise_id': exercise.id,
                    'type': exercise.type,
                    'question': exercise.question[:50] + "...",
                    'attempts': total_count,
                    'correct': correct_count,
                    'success_rate': success_rate
                })
            
            module_data.append({
                'module_id': module.id,
                'classic_text': module.classic_text[:100] + "...",
                'created_at': module.created_at,
                'total_attempts': total_attempts,
                'completed_attempts': completed_attempts,
                'completion_rate': (completed_attempts / total_attempts * 100) if total_attempts > 0 else 0,
                'avg_score': avg_score,
                'avg_accuracy': avg_accuracy,
                'exercise_count': len(exercises_list),
                'exercise_stats': exercise_stats
            })
        
        if format == 'pdf':
            return self._generate_pdf_report('module_performance', {'modules': module_data})
        elif format == 'excel':
            return self._generate_excel_report('module_performance', {'modules': module_data})
        else:
            return json.dumps({'modules': module_data}, default=str).encode()
    
    @db_session
    def generate_exercise_analysis_report(self, format: str = 'pdf') -> bytes:
        """Generate detailed analysis of all exercises"""
        
        exercise_types = [
            'multiple_choice', 'fill_in_blank', 'true_false', 'matching',
            'error_correction', 'transformation', 'ordering', 'completion'
        ]
        
        analysis_data = {}
        
        for ex_type in exercise_types:
            # ✅ FIX: Convert to list immediately
            exercises_list = list(Exercise.select(lambda e: e.type == ex_type))
            
            type_stats = {
                'total_exercises': len(exercises_list),
                'total_attempts': 0,
                'correct_answers': 0,
                'avg_success_rate': 0,
                'difficulty_distribution': {'easy': 0, 'medium': 0, 'hard': 0},
                'top_challenging': [],
                'top_easy': []
            }
            
            exercise_details = []
            
            for exercise in exercises_list:
                answers_list = list(exercise.user_answers)
                
                correct_count = sum(1 for a in answers_list if a.is_correct)
                total_count = len(answers_list)
                success_rate = (correct_count / total_count * 100) if total_count > 0 else 0
                
                type_stats['total_attempts'] += total_count
                type_stats['correct_answers'] += correct_count
                
                # Categorize difficulty
                if success_rate >= 80:
                    difficulty = 'easy'
                elif success_rate >= 50:
                    difficulty = 'medium'
                else:
                    difficulty = 'hard'
                
                type_stats['difficulty_distribution'][difficulty] += 1
                
                exercise_details.append({
                    'exercise_id': exercise.id,
                    'question': exercise.question,
                    'module_id': exercise.module.id,
                    'attempts': total_count,
                    'success_rate': success_rate,
                    'difficulty': difficulty
                })
            
            # Calculate average success rate
            if type_stats['total_attempts'] > 0:
                type_stats['avg_success_rate'] = (
                    type_stats['correct_answers'] / type_stats['total_attempts'] * 100
                )
            
            # Sort exercises by success rate
            exercise_details.sort(key=lambda x: x['success_rate'])
            
            # Get top challenging and easy exercises
            if len(exercise_details) >= 3:
                type_stats['top_challenging'] = exercise_details[:3]
                type_stats['top_easy'] = exercise_details[-3:]
            
            analysis_data[ex_type] = {
                'stats': type_stats,
                'exercises': exercise_details
            }
        
        if format == 'pdf':
            return self._generate_pdf_report('exercise_analysis', analysis_data)
        elif format == 'excel':
            return self._generate_excel_report('exercise_analysis', analysis_data)
        else:
            return json.dumps(analysis_data, default=str).encode()
    
    @db_session
    def generate_engagement_report(self, date_from: datetime = None, date_to: datetime = None, format: str = 'pdf') -> bytes:
        """Generate student engagement metrics report"""
        
        if not date_from:
            date_from = datetime.now() - timedelta(days=30)
        if not date_to:
            date_to = datetime.now()
        
        engagement_data = {
            'date_from': date_from.strftime('%Y-%m-%d'),
            'date_to': date_to.strftime('%Y-%m-%d'),
            'daily_active_users': self._get_daily_active_users(date_from, date_to),
            'module_completion_trends': self._get_completion_trends(date_from, date_to),
            'average_session_duration': self._get_avg_session_duration(date_from, date_to),
            'peak_usage_hours': self._get_peak_usage_hours(date_from, date_to),
            'student_retention': self._calculate_retention_rate(date_from, date_to),
            'engagement_by_module_type': self._get_engagement_by_content(date_from, date_to)
        }
        
        if format == 'pdf':
            return self._generate_pdf_report('engagement_metrics', engagement_data)
        elif format == 'excel':
            return self._generate_excel_report('engagement_metrics', engagement_data)
        else:
            return json.dumps(engagement_data, default=str).encode()
    
    # ========================================================================
    # HELPER METHODS - PYTHON 3.13 COMPATIBLE
    # ========================================================================
    
    @db_session
    def _get_exercise_performance(self, student_id: int) -> Dict[str, Any]:
        """Get exercise performance breakdown by type for a student"""
        
        user = User[student_id]
        exercise_types = [
            'multiple_choice', 'fill_in_blank', 'true_false', 'matching',
            'error_correction', 'transformation', 'ordering', 'completion'
        ]
        
        performance = {}
        
        for ex_type in exercise_types:
            # ✅ FIX: Convert to list immediately
            answers_list = []
            all_answers = list(UserAnswer.select())
            for a in all_answers:
                if a.progress.user.id == user.id and a.exercise.type == ex_type:
                    answers_list.append(a)
            
            if answers_list:
                correct = sum(1 for a in answers_list if a.is_correct)
                total = len(answers_list)
                performance[ex_type] = {
                    'attempts': total,
                    'correct': correct,
                    'accuracy': (correct / total * 100) if total > 0 else 0
                }
            else:
                performance[ex_type] = {
                    'attempts': 0,
                    'correct': 0,
                    'accuracy': 0
                }
        
        return performance
    
    @db_session
    def _get_weekly_progress(self, student_id: int, weeks: int = 12) -> List[Dict[str, Any]]:
        """Get weekly progress data for charts"""
        
        user = User[student_id]
        end_date = datetime.now()
        weekly_data = []
        
        for week in range(weeks):
            week_start = end_date - timedelta(days=(week + 1) * 7)
            week_end = end_date - timedelta(days=week * 7)
            
            # ✅ FIX: Get as list first
            all_progress = list(user.progress_records)
            week_progress = [
                p for p in all_progress 
                if p.started_at >= week_start and p.started_at < week_end
            ]
            
            accuracy_sum = 0
            for p in week_progress:
                if p.total_questions > 0:
                    accuracy_sum += (p.correct_answers / p.total_questions * 100)
            
            weekly_data.append({
                'week': week_end.strftime('%Y-%m-%d'),
                'modules_started': len(week_progress),
                'modules_completed': sum(1 for p in week_progress if p.completed),
                'total_score': sum(p.total_score for p in week_progress),
                'accuracy': accuracy_sum / len(week_progress) if week_progress else 0
            })
        
        weekly_data.reverse()
        return weekly_data
    
    @db_session
    def _get_daily_active_users(self, date_from: datetime, date_to: datetime) -> List[Dict]:
        """Calculate daily active users"""
        
        daily_data = []
        current_date = date_from
        
        # ✅ FIX: Get all progress first
        all_progress = list(UserProgress.select())
        
        while current_date <= date_to:
            next_date = current_date + timedelta(days=1)
            
            active_users = set()
            for prog in all_progress:
                if prog.started_at >= current_date and prog.started_at < next_date:
                    active_users.add(prog.user.id)
            
            daily_data.append({
                'date': current_date.strftime('%Y-%m-%d'),
                'active_users': len(active_users)
            })
            
            current_date = next_date
        
        return daily_data
    
    @db_session
    def _get_completion_trends(self, date_from: datetime, date_to: datetime) -> List[Dict]:
        """Get module completion trends"""
        
        weekly_trends = []
        current_week = date_from
        
        # ✅ FIX: Get all progress first
        all_progress = list(UserProgress.select())
        
        while current_week <= date_to:
            week_end = current_week + timedelta(days=7)
            
            completions = 0
            for p in all_progress:
                if p.completed and p.completed_at:
                    if p.completed_at >= current_week and p.completed_at < week_end:
                        completions += 1
            
            weekly_trends.append({
                'week': current_week.strftime('%Y-%m-%d'),
                'completions': completions
            })
            
            current_week = week_end
        
        return weekly_trends
    
    @db_session
    def _get_avg_session_duration(self, date_from: datetime, date_to: datetime) -> float:
        """Calculate average session duration in minutes"""
        
        # ✅ FIX: Get all progress first
        all_progress = list(UserProgress.select())
        
        total_duration = 0
        count = 0
        
        for prog in all_progress:
            if (prog.started_at >= date_from and prog.started_at <= date_to and 
                prog.completed_at and prog.started_at):
                duration = (prog.completed_at - prog.started_at).total_seconds() / 60
                total_duration += duration
                count += 1
        
        return round(total_duration / count, 1) if count > 0 else 0
    
    @db_session
    def _get_peak_usage_hours(self, date_from: datetime, date_to: datetime) -> List[Dict]:
        """Get peak usage hours distribution"""
        
        # ✅ FIX: Get all progress first
        all_progress = list(UserProgress.select())
        
        hour_counts = {}
        for prog in all_progress:
            if prog.started_at >= date_from and prog.started_at <= date_to:
                hour = prog.started_at.hour
                hour_counts[hour] = hour_counts.get(hour, 0) + 1
        
        hours_data = []
        for hour in range(24):
            hours_data.append({
                'hour': f"{hour:02d}:00",
                'activity_count': hour_counts.get(hour, 0)
            })
        
        hours_data.sort(key=lambda x: x['activity_count'], reverse=True)
        return hours_data[:5]
    
    @db_session
    def _get_engagement_by_content(self, date_from: datetime, date_to: datetime) -> Dict[str, Any]:
        """Get engagement metrics by content type"""
        
        exercise_types = {}
        
        # ✅ FIX: Get all exercises first
        all_exercises = list(Exercise.select())
        
        for exercise in all_exercises:
            ex_type = exercise.type
            if ex_type not in exercise_types:
                exercise_types[ex_type] = {
                    'total_attempts': 0,
                    'unique_users': set()
                }
            
            # Get answers in date range
            all_answers = list(exercise.user_answers)
            for answer in all_answers:
                if answer.answered_at >= date_from and answer.answered_at <= date_to:
                    exercise_types[ex_type]['total_attempts'] += 1
                    exercise_types[ex_type]['unique_users'].add(answer.progress.user.id)
        
        # Convert sets to counts
        result = {}
        for ex_type, data in exercise_types.items():
            result[ex_type] = {
                'total_attempts': data['total_attempts'],
                'unique_users': len(data['unique_users'])
            }
        
        return result
    
    @db_session
    def _calculate_retention_rate(self, date_from: datetime, date_to: datetime) -> float:
        """Calculate student retention rate"""
        
        all_students = list(User.select(lambda u: u.role == 'student'))
        total_students = len(all_students)
        
        active_students = 0
        for student in all_students:
            progress_list = list(student.progress_records)
            has_recent_activity = any(
                p.started_at >= date_from for p in progress_list
            )
            if has_recent_activity:
                active_students += 1
        
        retention_rate = (active_students / total_students * 100) if total_students > 0 else 0
        return round(retention_rate, 1)
    
    # ========================================================================
    # PDF GENERATION (Same as before - no changes needed)
    # ========================================================================
    
    def _generate_pdf_report(self, report_type: str, data: Dict[str, Any]) -> bytes:
        """Generate PDF report"""
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        title = self.REPORT_TYPES.get(report_type, 'Report')
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 20))
        
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_RIGHT
        )
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", date_style))
        story.append(Spacer(1, 20))
        
        if report_type == 'student_progress':
            self._add_student_progress_content(story, data, styles)
        elif report_type == 'class_overview':
            self._add_class_overview_content(story, data, styles)
        elif report_type == 'module_performance':
            self._add_module_performance_content(story, data, styles)
        elif report_type == 'exercise_analysis':
            self._add_exercise_analysis_content(story, data, styles)
        elif report_type == 'engagement_metrics':
            self._add_engagement_metrics_content(story, data, styles)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.read()
    
    def _add_student_progress_content(self, story: List, data: Dict, styles):
        """Add student progress content to PDF"""
        
        student = data['student']
        
        story.append(Paragraph(f"<b>Student:</b> {student.full_name} ({student.username})", styles['Normal']))
        story.append(Paragraph(f"<b>Email:</b> {student.email}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Modules Attempted', str(data['total_modules'])],
            ['Modules Completed', str(data['completed_modules'])],
            ['Total Score', str(data['total_score'])],
            ['Overall Accuracy', f"{data['accuracy']:.1f}%"]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        story.append(Paragraph("<b>Module Progress Details</b>", styles['Heading2']))
        story.append(Spacer(1, 10))
        
        module_data = [['Module', 'Status', 'Score', 'Accuracy', 'Date']]
        for module in data['module_details'][:10]:  # Limit to 10 for PDF
            module_data.append([
                module['module_title'],
                'Completed' if module['completed'] else 'In Progress',
                str(module['score']),
                f"{module['accuracy']:.1f}%",
                module['started_at'].strftime('%Y-%m-%d')
            ])
        
        module_table = Table(module_data, repeatRows=1)
        module_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')])
        ]))
        
        story.append(module_table)
    
    def _add_class_overview_content(self, story: List, data: Dict, styles):
        """Add class overview content to PDF"""
        
        story.append(Paragraph("<b>Class Statistics Summary</b>", styles['Heading2']))
        story.append(Spacer(1, 10))
        
        stats = data['class_stats']
        summary_data = [
            ['Metric', 'Value'],
            ['Total Students', str(stats.get('total_students', 0))],
            ['Active Students', str(stats.get('active_students', 0))],
            ['Average Score', f"{stats.get('avg_score', 0):.1f}"],
            ['Average Accuracy', f"{stats.get('avg_accuracy', 0):.1f}%"]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        story.append(Paragraph("<b>Top 10 Students</b>", styles['Heading2']))
        story.append(Spacer(1, 10))
        
        student_data = [['Rank', 'Student', 'Score', 'Modules', 'Accuracy']]
        for idx, student in enumerate(data['class_data'][:10], 1):
            student_data.append([
                str(idx),
                student['student_name'],
                str(student['total_score']),
                str(student['modules_completed']),
                f"{student['accuracy']:.1f}%"
            ])
        
        student_table = Table(student_data, repeatRows=1)
        student_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')])
        ]))
        
        story.append(student_table)
    
    def _add_module_performance_content(self, story: List, data: Dict, styles):
        """Add module performance content to PDF"""
        
        story.append(Paragraph("<b>Module Performance Analysis</b>", styles['Heading2']))
        story.append(Spacer(1, 10))
        
        for module in data['modules'][:5]:  # Limit to 5 modules for PDF
            story.append(Paragraph(f"<b>Module:</b> {module['classic_text']}", styles['Normal']))
            story.append(Spacer(1, 5))
            
            mod_stats = [
                ['Metric', 'Value'],
                ['Total Attempts', str(module['total_attempts'])],
                ['Completed', str(module['completed_attempts'])],
                ['Completion Rate', f"{module['completion_rate']:.1f}%"],
                ['Average Score', f"{module['avg_score']:.1f}"],
                ['Average Accuracy', f"{module['avg_accuracy']:.1f}%"]
            ]
            
            mod_table = Table(mod_stats)
            mod_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(mod_table)
            story.append(Spacer(1, 20))
    
    def _add_exercise_analysis_content(self, story: List, data: Dict, styles):
        """Add exercise analysis content to PDF"""
        
        story.append(Paragraph("<b>Exercise Type Analysis</b>", styles['Heading2']))
        story.append(Spacer(1, 10))
        
        for ex_type, type_data in list(data.items())[:5]:  # Limit to 5 types
            if not type_data.get('stats'):
                continue
                
            stats = type_data['stats']
            
            story.append(Paragraph(f"<b>{ex_type.replace('_', ' ').title()}</b>", styles['Heading3']))
            story.append(Spacer(1, 5))
            
            type_stats = [
                ['Metric', 'Value'],
                ['Total Exercises', str(stats['total_exercises'])],
                ['Total Attempts', str(stats['total_attempts'])],
                ['Success Rate', f"{stats['avg_success_rate']:.1f}%"],
                ['Easy', str(stats['difficulty_distribution']['easy'])],
                ['Medium', str(stats['difficulty_distribution']['medium'])],
                ['Hard', str(stats['difficulty_distribution']['hard'])]
            ]
            
            type_table = Table(type_stats)
            type_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(type_table)
            story.append(Spacer(1, 20))
    
    def _add_engagement_metrics_content(self, story: List, data: Dict, styles):
        """Add engagement metrics content to PDF"""
        
        story.append(Paragraph("<b>Student Engagement Metrics</b>", styles['Heading2']))
        story.append(Spacer(1, 10))
        
        if 'daily_active_users' in data:
            story.append(Paragraph("<b>Recent Daily Active Users</b>", styles['Heading3']))
            story.append(Spacer(1, 5))
            
            dau_data = [['Date', 'Active Users']]
            for day in data['daily_active_users'][-10:]:  # Last 10 days
                dau_data.append([day['date'], str(day['active_users'])])
            
            dau_table = Table(dau_data)
            dau_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(dau_table)
    
    # ========================================================================
    # EXCEL GENERATION (Simplified - same logic as PDF but for Excel)
    # ========================================================================
    
    def _generate_excel_report(self, report_type: str, data: Dict[str, Any]) -> bytes:
        """Generate Excel report"""
        
        wb = Workbook()
        wb.remove(wb.active)
        
        if report_type == 'student_progress':
            self._add_student_progress_sheets(wb, data)
        elif report_type == 'class_overview':
            self._add_class_overview_sheets(wb, data)
        elif report_type == 'module_performance':
            self._add_module_performance_sheets(wb, data)
        elif report_type == 'exercise_analysis':
            self._add_exercise_analysis_sheets(wb, data)
        elif report_type == 'engagement_metrics':
            self._add_engagement_metrics_sheets(wb, data)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
    
    def _add_student_progress_sheets(self, wb: Workbook, data: Dict):
        """Add student progress data to Excel"""
        
        ws = wb.create_sheet("Student Progress")
        ws.append(["Student Progress Report"])
        ws.append([])
        ws.append(["Student", data['student'].full_name])
        ws.append(["Total Modules", len(data['module_details'])])
        ws.append(["Completed", sum(1 for m in data['module_details'] if m['completed'])])
        
        ws_modules = wb.create_sheet("Modules")
        ws_modules.append(["Module", "Status", "Score", "Accuracy", "Date"])
        for m in data['module_details']:
            ws_modules.append([
                m['module_title'],
                'Completed' if m['completed'] else 'In Progress',
                m['score'],
                f"{m['accuracy']:.1f}%",
                m['started_at'].strftime('%Y-%m-%d')
            ])
        
        self._format_excel_sheet(ws)
        self._format_excel_sheet(ws_modules)
    
    def _add_class_overview_sheets(self, wb: Workbook, data: Dict):
        """Add class overview to Excel"""
        
        ws = wb.create_sheet("Class Overview")
        ws.append(["Student", "Score", "Modules", "Accuracy"])
        for s in data['class_data']:
            ws.append([
                s['student_name'],
                s['total_score'],
                s['modules_completed'],
                f"{s['accuracy']:.1f}%"
            ])
        
        self._format_excel_sheet(ws)
    
    def _add_module_performance_sheets(self, wb: Workbook, data: Dict):
        """Add module performance to Excel"""
        
        ws = wb.create_sheet("Modules")
        ws.append(["Module", "Attempts", "Completed", "Completion %", "Avg Score", "Avg Accuracy"])
        for m in data['modules']:
            ws.append([
                m['classic_text'],
                m['total_attempts'],
                m['completed_attempts'],
                f"{m['completion_rate']:.1f}%",
                f"{m['avg_score']:.1f}",
                f"{m['avg_accuracy']:.1f}%"
            ])
        
        self._format_excel_sheet(ws)
    
    def _add_exercise_analysis_sheets(self, wb: Workbook, data: Dict):
        """Add exercise analysis to Excel"""
        
        for ex_type, type_data in list(data.items())[:3]:  # Limit
            if not type_data.get('stats'):
                continue
            
            ws = wb.create_sheet(ex_type[:30])
            stats = type_data['stats']
            ws.append([ex_type, "Stats"])
            ws.append(["Exercises", stats['total_exercises']])
            ws.append(["Attempts", stats['total_attempts']])
            ws.append(["Success Rate", f"{stats['avg_success_rate']:.1f}%"])
            
            self._format_excel_sheet(ws)
    
    def _add_engagement_metrics_sheets(self, wb: Workbook, data: Dict):
        """Add engagement metrics to Excel"""
        
        if 'daily_active_users' in data:
            ws = wb.create_sheet("Daily Active Users")
            ws.append(["Date", "Active Users"])
            for d in data['daily_active_users']:
                ws.append([d['date'], d['active_users']])
            
            self._format_excel_sheet(ws)
    
    def _format_excel_sheet(self, ws):
        """Format Excel worksheet"""
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def generate_pdf_report(self, report_data: dict) -> bytes:
        """Generate PDF report from report data"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24,
            textColor=colors.HexColor('#667eea'), spaceAfter=30, alignment=TA_CENTER)
        heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=16,
            textColor=colors.HexColor('#333333'), spaceAfter=12)
        
        # Title
        report_type = report_data.get('report_type', 'Report')
        elements.append(Paragraph(report_type, title_style))
        
        # Generated date
        generated_at = report_data.get('generated_at', datetime.now().isoformat())
        date_text = f"Generated: {generated_at[:19].replace('T', ' ')}"
        elements.append(Paragraph(date_text, styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Summary section
        if 'summary' in report_data:
            elements.append(Paragraph("Summary", heading_style))
            summary_data = []
            for key, value in report_data['summary'].items():
                if value is not None:
                    label = key.replace('_', ' ').title()
                    summary_data.append([label, str(value)])
            
            if summary_data:
                summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8)
                ]))
                elements.append(summary_table)
                elements.append(Spacer(1, 20))
        
        # Data section
        if 'data' in report_data and report_data['data']:
            elements.append(Paragraph("Detailed Data", heading_style))
            data_list = report_data['data']
            
            if isinstance(data_list, list) and len(data_list) > 0:
                first_item = data_list[0]
                headers = list(first_item.keys())
                table_data = [headers]
                
                for item in data_list:
                    row = []
                    for header in headers:
                        value = item.get(header, '')
                        if isinstance(value, (list, dict)):
                            value = str(len(value)) if isinstance(value, list) else 'Object'
                        row.append(str(value)[:50])
                    table_data.append(row)
                
                col_width = (doc.width) / len(headers)
                data_table = Table(table_data, colWidths=[col_width] * len(headers))
                data_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6)
                ]))
                elements.append(data_table)
        
        # Students section
        elif 'students' in report_data and report_data['students']:
            elements.append(Paragraph("Students", heading_style))
            students_list = report_data['students']
            
            if isinstance(students_list, list) and len(students_list) > 0:
                first_student = students_list[0]
                headers = list(first_student.keys())
                table_data = [headers]
                
                for student in students_list:
                    row = []
                    for header in headers:
                        value = student.get(header, '')
                        if isinstance(value, list):
                            value = ', '.join(str(v) for v in value[:3])
                        elif isinstance(value, dict):
                            value = 'Object'
                        row.append(str(value)[:50])
                    table_data.append(row)
                
                col_width = (doc.width) / len(headers)
                students_table = Table(table_data, colWidths=[col_width] * len(headers))
                students_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6)
                ]))
                elements.append(students_table)
        
        # Weekly summary
        elif 'weekly_summary' in report_data:
            elements.append(Paragraph("Weekly Summary", heading_style))
            weekly_data = report_data['weekly_summary']
            summary_items = []
            
            for key, value in weekly_data.items():
                label = key.replace('_', ' ').title()
                summary_items.append([label, str(value)])
            
            weekly_table = Table(summary_items, colWidths=[3*inch, 3*inch])
            weekly_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8)
            ]))
            elements.append(weekly_table)
        
        doc.build(elements)
        pdf_content = buffer.getvalue()
        buffer.close()
        return pdf_content
    
    def generate_excel_report(self, report_data: dict) -> bytes:
        """Generate Excel report from report data"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        row = 1
        
        # Title
        report_type = report_data.get('report_type', 'Report')
        ws.merge_cells(f'A{row}:F{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = report_type
        title_cell.font = Font(name='Arial', size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        row += 2
        
        # Generated date
        ws[f'A{row}'] = 'Generated:'
        ws[f'B{row}'] = report_data.get('generated_at', datetime.now().isoformat())[:19]
        row += 2
        
        # Summary section
        if 'summary' in report_data:
            ws[f'A{row}'] = 'Summary'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            for key, value in report_data['summary'].items():
                if value is not None:
                    ws[f'A{row}'] = key.replace('_', ' ').title()
                    ws[f'B{row}'] = str(value)
                    row += 1
            row += 1
        
        # Data section
        if 'data' in report_data and report_data['data']:
            data_list = report_data['data']
            
            if isinstance(data_list, list) and len(data_list) > 0:
                headers = list(data_list[0].keys())
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header.replace('_', ' ').title()
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                row += 1
                
                for item in data_list:
                    for col, header in enumerate(headers, start=1):
                        value = item.get(header, '')
                        if isinstance(value, list):
                            value = ', '.join(str(v) for v in value[:3])
                        elif isinstance(value, dict):
                            value = 'Object'
                        cell = ws.cell(row=row, column=col)
                        cell.value = str(value)
                        cell.border = border
                    row += 1
        
        # Students section
        elif 'students' in report_data and report_data['students']:
            students_list = report_data['students']
            
            if isinstance(students_list, list) and len(students_list) > 0:
                headers = list(students_list[0].keys())
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header.replace('_', ' ').title()
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                row += 1
                
                for student in students_list:
                    for col, header in enumerate(headers, start=1):
                        value = student.get(header, '')
                        if isinstance(value, list):
                            value = ', '.join(str(v) for v in value[:3])
                        elif isinstance(value, dict):
                            value = 'Object'
                        cell = ws.cell(row=row, column=col)
                        cell.value = str(value)
                        cell.border = border
                    row += 1
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
        
            for cell in column:
            # ✅ Skip merged cells
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
            
            # Get column letter from first non-merged cell
                if column_letter is None:
                    column_letter = cell.column_letter
            
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:  
                    pass
        
        # ✅ Apply width only if we found a valid column letter
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = BytesIO()
        wb.save(buffer)
        excel_content = buffer.getvalue()
        buffer.close()
        return excel_content


report_service = ReportService()